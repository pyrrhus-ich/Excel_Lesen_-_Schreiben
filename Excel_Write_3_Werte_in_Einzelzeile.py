from openpyxl import Workbook
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
myList=['Frank','hatte','gestern','Geburtstag']
spalten=[]
    #  befüllt die Liste spalten mit der benötigten Anzahl Buchstaben
for x in list(range(1,len(myList)+1)):
    spaltenName = get_column_letter(x)
    spalten.append(spaltenName)
    # schreibt die Werte aus myList in das Excel. Wobei die Zeile 
    # immer gleich bleibt 'zeilenNummer' und nur der Spaltenbuch-
    # stabe sich ändert
for el in spalten:
    i=0
    zeilenNummer = "1"
    while i < len(spalten):
        ws[spalten[i]+zeilenNummer]=myList[i]
        i+=1

wb.save("Bsp_Excel_Write_3.xlsx")




