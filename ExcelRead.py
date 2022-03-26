from openpyxl import load_workbook

#Basis für Nutzung von openpyxl 
# 1. Workbook erstellen => hier "wb"
wb=load_workbook(filename="Test.xlsx")
# 2. Worksheet erstellen => hier ws
ws=wb["Tabelle1"]
# Iterator erstellen
def iter_rows(ws):
    for row in ws.iter_rows():
        yield [cell.value for cell in row]
print(list(iter_rows(ws)))

val = []
for el in iter_rows(ws):
    val.append(el)
print(val)
val0=[]
for el in iter_rows(ws):
    val0.append(el[0])
print(val0)





